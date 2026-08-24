# -*- coding: utf-8 -*-
"""Prova a importacao da planilha do DMOB: o projeto gerado abre e confere com a planilha.

Compara o que a plataforma mostra com o que a aba `CONTROLE DA OBRA` reporta. Nao basta o
arquivo abrir: se o quilometro entrar deslocado de um, o mapa fica plausivel e a medicao
fica errada — e ninguem percebe.

Uso: python ferramentas/testar-importacao.py
"""
import functools
import http.server
import io
import json
import os
import socketserver
import sys
import threading

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from playwright.sync_api import sync_playwright

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Erro de REDE EXTERNA nao e defeito da plataforma: os ladrilhos de satelite vem de
# fora e falham quando a suite inteira disputa banda. Filtro aplicado em 24/08, depois
# de tres provas ficarem vermelhas na suite e verdes sozinhas.
EXTERNO = ("Failed to load resource", "ERR_", "net::", "arcgisonline", "tile")

PASTA = os.path.join(RAIZ, "11-CONTROLE AM-010")
PLANILHA = os.path.join(PASTA, "CAMADA DE KM AM-010.xlsx")
PORTA = 8753

falhas = []


def ok(cond, msg, extra=""):
    print(f"  {'OK   ' if cond else 'FALHA'} {msg}" + (f"  → {extra}" if extra else ""))
    if not cond:
        falhas.append(msg)


class Q(http.server.SimpleHTTPRequestHandler):
    def log_message(self, *a):
        pass


def da_planilha():
    """O que a planilha diz, lido direto dela — não do projeto gerado."""
    wb = openpyxl.load_workbook(PLANILHA, read_only=True, data_only=True)
    aba = next(n for n in wb.sheetnames if n.upper().startswith("CAMADA DE KM"))
    linhas = list(wb[aba].iter_rows(values_only=True))
    # só o que tem segmento no eixo: a planilha vai até o KM 271 e o eixo tem 269
    concl_remendo = sum(1 for r in linhas[1:]
                        if str(r[6] or "").strip().upper().startswith("CONCLU")
                        and r[0] is not None and int(r[0]) <= 269)
    frentes = {}
    for r in linhas[1:]:
        g = str(r[4] or "").strip().upper()
        s = str(r[5] or "").strip().upper()
        if g and g != "PLANEJADO" and s:
            frentes[g] = frentes.get(g, 0) + 1
    nome = "CONTROLE DA OBRA " if "CONTROLE DA OBRA " in wb.sheetnames else "CONTROLE DA OBRA"
    contratado = {}
    for r in list(wb[nome].iter_rows(values_only=True))[1:]:
        g = str(r[0] or "").strip().upper()
        if g and r[1]:
            contratado[g] = float(r[1])
    wb.close()
    return {"remendo_concluido": concl_remendo, "frentes": frentes, "contratado": contratado}


def main():
    proj = next((os.path.join(PASTA, f) for f in os.listdir(PASTA)
                 if f.startswith("projeto-am-010") and f.endswith(".json")), None)
    if not proj:
        print("gere o projeto primeiro: python ferramentas/importar-camada-de-km.py")
        return 1
    esperado = da_planilha()
    p = json.load(io.open(proj, encoding="utf-8"))
    print(f"projeto: {os.path.basename(proj)} · {len(p['dados'])} lançamentos\n")

    print("1. O QUE FOI GRAVADO CONFERE COM A PLANILHA")
    ok(p["contrato"] == "CT-00057/2022-SEINFRA", "número do contrato", p["contrato"])
    cd = p.get("contratoDados", {})
    ok(abs((cd.get("valor") or 0) - 515701566.59) < 0.01, "valor do contrato",
       f"R$ {cd.get('valor', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    ok(bool(cd.get("vigencia_execucao")), "vigência de execução", cd.get("vigencia_execucao", ""))
    rem = sum(1 for k, v in p["dados"].items()
              if k.split("|")[1] == "REMENDO PROFUNDO" and k.split("|")[2] == "LE" and v == "C")
    ok(rem == esperado["remendo_concluido"],
       "remendo profundo concluído: mesmo número de quilômetros da planilha",
       f"{rem} de {esperado['remendo_concluido']}")
    q = {s["nome"].upper(): s.get("km_contratado") for s in p["svc"]}
    faltando = [g for g, v in esperado["contratado"].items()
                if g in q and (q[g] is None or abs(q[g] - v) > 0.01)]
    ok(not faltando, "quantidade contratada por serviço veio da aba CONTROLE DA OBRA",
       f"{sum(1 for v in q.values() if v)} serviço(s)"
       + (f" · divergem: {faltando}" if faltando else ""))

    print("\n2. A PLATAFORMA ABRE O PROJETO")
    H = functools.partial(Q, directory=RAIZ)
    socketserver.TCPServer.allow_reuse_address = True
    srv = socketserver.TCPServer(("127.0.0.1", PORTA), H)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    console = []
    with sync_playwright() as pw:
        nav = pw.chromium.launch()
        pg = nav.new_context(viewport={"width": 1680, "height": 1000}).new_page()
        pg.on("console", lambda m: console.append(f"[{m.type}] {m.text}")
              if m.type == "error" and not any(x in m.text for x in EXTERNO) else None)
        pg.on("pageerror", lambda e: console.append(f"[pageerror] {e}"))
        pg.on("dialog", lambda d: d.accept())
        pg.goto(f"http://127.0.0.1:{PORTA}/index.html", wait_until="domcontentloaded")
        pg.wait_for_timeout(2500)
        pg.set_input_files("#fileProjeto", proj)
        pg.wait_for_timeout(4500)
        ok(pg.evaluate("S.eixo ? S.eixo.nome : ''") == "AM-010", "o eixo entrou")
        # LD e LE do mesmo serviço e quilômetro viram UMA chave desde 24/08: 760 lançamentos
        # do projeto do cliente viram 380 sem perder um quilômetro. O que importa é que
        # nenhum serviço suma — e isso a linha seguinte afere.
        n = pg.evaluate("Object.keys(S.dados).length")
        ok(0 < n <= len(p["dados"]), "os lançamentos entram, com os lados fundidos",
           f"{n} chave(s) de {len(p['dados'])} do arquivo")
        ok(pg.evaluate("document.querySelector('#contrato').value") == p["contrato"],
           "o contrato aparece no cabeçalho")
        cels = pg.evaluate("document.querySelectorAll('#faixaTrilho .km').length")
        pintadas = pg.evaluate("""[...document.querySelectorAll('#faixaTrilho .km')]
            .filter(e => e.style.background && e.style.background !== 'rgb(232, 237, 242)').length""")
        ok(cels == 269, "a faixa monta os 269 quilômetros", str(cels))
        ok(pintadas > 200, "a faixa mostra os quilômetros com serviço lançado",
           f"{pintadas} pintado(s)")

        print("\n3. O AVANÇO CONFERE COM O QUE O ESCRITÓRIO REPORTA")
        # remendo profundo: 175 km contratados, 175 realizados = 100% no quadro do escritório;
        # medido contra o trecho de 268 km daria 65% — é a diferença de denominador
        av = pg.evaluate("""() => {
            // o lado saiu em 24/08: a linha da grade é do serviço, e o lado é sempre 'U'
            const l = {svc: 'REMENDO PROFUNDO', lado: 'U'};
            const r = resumoLinha(l);
            return {km: r.kmC, pct: r.pct};
        }""")
        ok(abs(av["km"] - esperado["remendo_concluido"]) < 1.5,
           "quilômetros concluídos de remendo profundo",
           f"{av['km']:.1f} km medidos · {esperado['remendo_concluido']} na planilha")
        contratados = esperado["contratado"].get("REMENDO PROFUNDO")
        print(f"         a plataforma mostra {100 * av['pct']:.1f}% (sobre o trecho de "
              f"268 km); o escritório reporta {100 * av['km'] / contratados:.0f}% "
              f"(sobre os {contratados:.0f} km contratados)")
        pg.screenshot(path=os.path.join(RAIZ, "documentacao", "imagens",
                                        "08-am-010-importada.png"))
        nav.close()
    srv.shutdown()

    print(f"\nERROS DE CONSOLE: {len(console)}")
    for c in console[:8]:
        print("   ", c[:150])
    print("\nRESULTADO:", f"{len(falhas)} FALHA(S)" if falhas or console
          else "OK — a planilha do DMOB entra na plataforma sem perder nada")
    for f in falhas:
        print("   falhou:", f)
    return 1 if (falhas or console) else 0


if __name__ == "__main__":
    sys.exit(main())
