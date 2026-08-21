# -*- coding: utf-8 -*-
"""Importa a planilha de controle do DMOB para um projeto do SICOR.

FONTE: `11-CONTROLE AM-010\\CAMADA DE KM AM-010.xlsx`, a planilha viva da obra da AM-010.
Tem tres abas que interessam:

- `CAMADA DE KM AM-010` — uma linha por quilometro, com GRUPO DE SERVICO (a frente de
  trabalho naquele km), SITUACAO GRUPO DE SERVICO e uma coluna por servico e lado (E/D).
- `CONTROLE DA OBRA` — o quadro que o escritorio reporta: por servico, o TOTAL (KM)
  CONTRATADO e quantos quilometros estao em cada situacao. O denominador do avanco e a
  quantidade contratada, nao a extensao do trecho: remendo profundo tem 175 km contratados
  numa obra de 250 km, e 175 realizados sao 100%, nao 70%.
- `DADOS CONTRATUAIS` — objeto, numero do contrato, valor e vigencias.

O QUE ESTE SCRIPT NAO FAZ: casar por coordenada. As coordenadas da planilha estao
arredondadas em tres casas e chegam a 1,4 km do eixo do acervo — casar por elas trocaria
quilometro de lugar. O casamento e pelo NUMERO do quilometro: KM 1 da planilha e o primeiro
segmento do eixo (KM 0 a 1).

Uso:  python ferramentas/importar-camada-de-km.py
      python ferramentas/importar-camada-de-km.py "caminho\\da\\planilha.xlsx"
"""
import io
import json
import math
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PADRAO = os.path.join(RAIZ, "11-CONTROLE AM-010", "CAMADA DE KM AM-010.xlsx")

# Vocabulario do escritorio -> codigo de situacao do SICOR. O que nao estiver aqui e
# reportado, nunca adivinhado: situacao errada num quilometro e medicao errada.
SITUACAO = {
    "CONCLUÍDO": "C", "CONCLUIDO": "C",
    "EM ANDAMENTO": "E",
    "PARALISADO": "PA", "PARADO": "PA",
    "PLANEJADO": "", "A INICIAR": "", "PREVISTO": "",
    "SEM PLANEJAMENTO": "S",
    "NÃO SE APLICA": "NA", "NAO SE APLICA": "NA",
}
# GRUPO DE SERVICO nomeia o servico da frente; estes dois nao sao servico
GRUPO_IGNORA = {"PLANEJADO", "SEM PLANEJAMENTO", ""}
# nome do grupo -> nome da coluna de servico, quando divergem
GRUPO_PARA_SVC = {"TRATAMENTO DE EROSÃO": "EROSÕES", "TRATAMENTO DE EROSAO": "EROSÕES"}
LADOS = {"E": "LE", "D": "LD"}


def limpa(v):
    return "" if v is None else str(v).strip()


def num(v):
    try:
        return float(str(v).replace(".", "").replace(",", ".")) if isinstance(v, str) else float(v)
    except (TypeError, ValueError):
        return None


def cabecalho_servico(txt):
    """«BASE E» -> ('BASE', 'LE'). «CONCRETO ASFÁLTICO» sem sufixo é o lado esquerdo: a
    planilha perdeu o « E» nessa coluna, e a coluna « D» ao lado prova qual é qual."""
    t = limpa(txt)
    if len(t) > 2 and t[-2] == " " and t[-1] in LADOS:
        return t[:-2].strip(), LADOS[t[-1]]
    return t, None


def main():
    arq = sys.argv[1] if len(sys.argv) > 1 else PADRAO
    if not os.path.exists(arq):
        print("não encontrei a planilha:", arq)
        return 1
    print("planilha:", os.path.basename(arq))
    wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)

    # ---------------------------------------------------------------- contrato
    contrato = {}
    if "DADOS CONTRATUAIS" in wb.sheetnames:
        r = list(wb["DADOS CONTRATUAIS"].iter_rows(values_only=True))
        if len(r) > 1:
            cab = [limpa(c).upper() for c in r[0]]
            val = list(r[1])
            de = lambda k: next((val[i] for i, c in enumerate(cab) if c.startswith(k)), None)
            contrato = {
                "objeto": limpa(de("OBJETO")),
                "numero": limpa(de("N°CT") or de("Nº CT") or de("N CT")),
                "valor": num(de("VALOR")),
                "vigencia_contrato": limpa(de("VIGENCIA DE CONTRATO") or de("VIGÊNCIA DE CONTRATO")),
                "vigencia_execucao": limpa(de("VIGENCIA DE EXEC") or de("VIGÊNCIA DE EXEC")),
            }
            print(f"contrato: {contrato['numero']} · "
                  f"R$ {contrato['valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    # ---------------------------------------------------------------- quantidade contratada
    contratado, ctrl_linhas, abas = {}, [], set(wb.sheetnames)
    if "CONTROLE DA OBRA " in wb.sheetnames or "CONTROLE DA OBRA" in wb.sheetnames:
        nome = "CONTROLE DA OBRA " if "CONTROLE DA OBRA " in wb.sheetnames else "CONTROLE DA OBRA"
        ctrl_linhas = list(wb[nome].iter_rows(values_only=True))
        for r in ctrl_linhas[1:]:
            g, t = limpa(r[0]).upper(), num(r[1])
            if g and t:
                contratado[GRUPO_PARA_SVC.get(g, g)] = t
        print(f"quantidade contratada: {len(contratado)} serviço(s) · "
              f"{sum(contratado.values()):,.0f} km somados".replace(",", "."))

    # ---------------------------------------------------------------- camada de km
    aba = next((n for n in wb.sheetnames if n.upper().startswith("CAMADA DE KM")), None)
    if not aba:
        print("a planilha não tem aba «CAMADA DE KM»")
        return 1
    linhas = list(wb[aba].iter_rows(values_only=True))
    wb.close()
    cab = [limpa(c) for c in linhas[0]]
    i_km = 0
    i_grupo = next((i for i, c in enumerate(cab) if c.upper() == "GRUPO DE SERVIÇO"), None)
    i_sit = next((i for i, c in enumerate(cab) if c.upper().startswith("SITUAÇÃO GRUPO")), None)
    colunas = []
    for i, c in enumerate(cab):
        if i < 4 or i in (i_grupo, i_sit) or not c or c.upper() == "PREVISÃO":
            continue
        svc, lado = cabecalho_servico(c)
        colunas.append((i, svc, lado or "LE"))
    print(f"colunas de serviço: {len(colunas)} · "
          f"{len({s for _, s, _ in colunas})} serviço(s) em dois lados")

    # ---------------------------------------------------------------- eixo do acervo
    acv = json.load(io.open(os.path.join(RAIZ, "dados", "acervo-rodovias-estaduais.json"),
                            encoding="utf-8"))
    nome_eixo = "AM-010"
    eixo = next((i for i in acv["itens"] if i["nome"] == nome_eixo), None)
    if not eixo:
        print("não achei", nome_eixo, "no acervo")
        return 1
    n_seg = contaSegmentos(eixo["linhas"])
    print(f"eixo: {nome_eixo} · {eixo['km_geometria']:.3f} km · {n_seg} segmentos")

    # ---------------------------------------------------------------- lançamentos
    dados, fora, sem_mapa, frentes = {}, 0, {}, 0
    for r in linhas[1:]:
        km = num(r[i_km])
        if km is None:
            continue
        seg = int(km) - 1                      # KM 1 da planilha = primeiro segmento (0)
        if seg < 0 or seg >= n_seg:
            fora += 1
            continue
        # a frente de trabalho do quilometro: o grupo diz o servico, a situacao diz o pe
        if i_grupo is not None and i_sit is not None:
            g = limpa(r[i_grupo]).upper()
            s = limpa(r[i_sit]).upper()
            if g not in GRUPO_IGNORA and s:
                svc = GRUPO_PARA_SVC.get(g, g)
                if s in SITUACAO:
                    cod = SITUACAO[s]
                    if cod:
                        for ld in ("LE", "LD"):
                            dados[f"recuperacao|{svc}|{ld}|{seg}"] = cod
                        frentes += 1
                else:
                    sem_mapa[s] = sem_mapa.get(s, 0) + 1
        # as colunas por servico e lado
        for i, svc, lado in colunas:
            v = limpa(r[i]).upper()
            if not v:
                continue
            if v not in SITUACAO:
                sem_mapa[v] = sem_mapa.get(v, 0) + 1
                continue
            cod = SITUACAO[v]
            if cod:
                dados[f"recuperacao|{svc}|{lado}|{seg}"] = cod

    print(f"lançamentos: {len(dados)} · frentes de serviço: {frentes}"
          + (f" · {fora} quilômetro(s) da planilha sem segmento no eixo" if fora else ""))
    if sem_mapa:
        print("  ATENÇÃO — situações que este script não sabe traduzir (ficaram de fora):")
        for k, v in sorted(sem_mapa.items(), key=lambda x: -x[1]):
            print(f"    «{k}» em {v} célula(s)")

    # ------------------------------------------------- conferência entre as duas abas
    # A `CAMADA DE KM` e a `CONTROLE DA OBRA` sao preenchidas por maos diferentes e podem
    # divergir. Medido em 21/08 na AM-010: a camada tem 259 km de remendo profundo marcados
    # como concluidos, e o quadro do escritorio reporta 175 realizados de 175 contratados.
    # Escolher uma das duas seria inventar numero de medicao; o importador denuncia.
    if contratado:
        realizado = {}
        nome_ctrl = ("CONTROLE DA OBRA " if "CONTROLE DA OBRA " in abas
                     else "CONTROLE DA OBRA" if "CONTROLE DA OBRA" in abas else None)
        if nome_ctrl:
            for r in ctrl_linhas[1:]:
                g, rz = limpa(r[0]).upper(), num(r[5])
                if g and rz is not None:
                    realizado[GRUPO_PARA_SVC.get(g, g)] = rz
        importado = {}
        for k, v in dados.items():
            _, svcN, ld, _seg = k.split("|")
            if ld == "LE" and v == "C":
                importado[svcN.upper()] = importado.get(svcN.upper(), 0) + 1
        diverge = []
        for g, rz in realizado.items():
            imp = importado.get(g, 0)
            if abs(imp - rz) > 1.5:
                diverge.append((g, imp, rz, contratado.get(g)))
        if diverge:
            print("\n  DIVERGÊNCIA ENTRE AS ABAS DA PLANILHA — nada foi corrigido aqui:")
            print("    serviço                          camada de km   controle da obra   contratado")
            for g, imp, rz, ct in sorted(diverge, key=lambda x: -abs(x[1] - x[2])):
                print(f"    {g[:32]:34s} {imp:8.0f} km   {rz:12.0f} km"
                      + (f"   {ct:7.0f} km" if ct else "        —"))
            print("    A plataforma importa o que está na CAMADA DE KM, quilômetro a "
                  "quilômetro, porque é ela que tem posição." + "\n"
                  + "    O quadro do escritório fica como referência.")

    # ---------------------------------------------------------------- serviços do projeto
    cat = json.load(io.open(os.path.join(RAIZ, "dados", "catalogo-servicos.json"),
                            encoding="utf-8"))
    porNome = {}
    for cj in cat["conjuntos"]:
        for s in cj["servicos"]:
            porNome[s["nome"].upper()] = s
    svc = []
    for nome in dict.fromkeys(s for _, s, _ in colunas):
        base = porNome.get(nome.upper(), {})
        svc.append({
            "nome": nome, "grupo": base.get("grupo", "Pavimentação"),
            "lados": ["LE", "LD"], "unidade": base.get("unidade", "m²"),
            "cor": base.get("cor", ""),
            "km_contratado": contratado.get(nome.upper()),
            "on": any(k.split("|")[1] == nome for k in dados),
        })
    print(f"serviços no projeto: {len(svc)} · "
          f"{sum(1 for s in svc if s['on'])} com lançamento · "
          f"{sum(1 for s in svc if s['km_contratado'])} com quantidade contratada")

    # ---------------------------------------------------------------- projeto
    proj = {
        "versao": 1,
        "obra": f"{nome_eixo} — {contrato.get('objeto', 'obra')}"[:120],
        "contrato": (contrato.get("numero") or "").upper(),
        "contratoDados": contrato,
        "ref": "km", "fonte": "rodovia", "catId": "recuperacao",
        "kmIni": 0, "kmFim": round(eixo["km_geometria"], 3),
        "estOff": 0, "invertido": False,
        "eixo": {
            "nome": eixo["nome"], "tipo": "rodovia", "linhas": eixo["linhas"],
            "km_cadastro": eixo["km_cadastro"], "km_geometria": eixo["km_geometria"],
            "inicio": eixo["inicio"], "fim": eixo["fim"],
            "sentido": eixo["sentido"], "meta": {"saltos_km": eixo["saltos_km"],
                                                 "partes": eixo["partes"]},
        },
        "svc": svc, "dados": dados,
        "ens": [], "reg": [], "seqReg": 0, "fotos": {},
        "origem": {
            "planilha": os.path.basename(arq),
            "aba": aba,
            "observacao": ("Importado da planilha de controle do DMOB. O casamento é pelo "
                           "número do quilômetro, não por coordenada: as coordenadas da "
                           "planilha estão arredondadas em três casas e chegam a 1,4 km do "
                           "eixo do acervo."),
        },
    }
    destino = os.path.join(os.path.dirname(arq),
                           f"projeto-{nome_eixo.lower()}-"
                           + (contrato.get("numero") or "sem-contrato")
                           .lower().replace("/", "-") + ".json")
    with io.open(destino, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(proj, fh, ensure_ascii=False, separators=(",", ":"))
    print(f"\ngravado: {os.path.relpath(destino, RAIZ)}"
          f" ({os.path.getsize(destino) / 1024:,.0f} KB)".replace(",", "."))
    print("abra no SICOR pelo botão «Abrir».")
    return 0


def contaSegmentos(linhas):
    """Quantos segmentos de 1 km o motor vai produzir — a mesma conta do produto."""
    A, F = 6378137.0, 1 / 298.257222101
    def geod(p1, p2):
        L = math.radians(p2[0] - p1[0])
        U1 = math.atan((1 - F) * math.tan(math.radians(p1[1])))
        U2 = math.atan((1 - F) * math.tan(math.radians(p2[1])))
        sU1, cU1, sU2, cU2 = math.sin(U1), math.cos(U1), math.sin(U2), math.cos(U2)
        lam, lamP, it = L, 2 * math.pi, 0
        while abs(lam - lamP) > 1e-12 and it < 40:
            sL, cL = math.sin(lam), math.cos(lam)
            sS = math.sqrt((cU2 * sL) ** 2 + (cU1 * sU2 - sU1 * cU2 * cL) ** 2)
            if sS == 0:
                return 0.0
            cS = sU1 * sU2 + cU1 * cU2 * cL
            sig = math.atan2(sS, cS)
            sA = cU1 * cU2 * sL / sS
            c2A = 1 - sA ** 2
            c2SM = 0 if c2A == 0 else cS - 2 * sU1 * sU2 / c2A
            C = F / 16 * c2A * (4 + F * (4 - 3 * c2A))
            lamP = lam
            lam = L + (1 - C) * F * sA * (sig + C * sS *
                  (c2SM + C * cS * (-1 + 2 * c2SM ** 2)))
            it += 1
        u2 = c2A * (A ** 2 - (A * (1 - F)) ** 2) / (A * (1 - F)) ** 2
        Ac = 1 + u2 / 16384 * (4096 + u2 * (-768 + u2 * (320 - 175 * u2)))
        B = u2 / 1024 * (256 + u2 * (-128 + u2 * (74 - 47 * u2)))
        sS = math.sqrt((math.cos(U2) * math.sin(lam)) ** 2 +
                       (cU1 * sU2 - sU1 * cU2 * math.cos(lam)) ** 2)
        cS = sU1 * sU2 + cU1 * cU2 * math.cos(lam)
        sig = math.atan2(sS, cS)
        c2SM = 0 if c2A == 0 else cS - 2 * sU1 * sU2 / c2A
        dS = B * sS * (c2SM + B / 4 * (cS * (-1 + 2 * c2SM ** 2) -
             B / 6 * c2SM * (-3 + 4 * sS ** 2) * (-3 + 4 * c2SM ** 2)))
        return (A * (1 - F)) * Ac * (sig - dS)
    total = 0.0
    for l in linhas:
        for i in range(len(l) - 1):
            total += geod(l[i], l[i + 1])
    return math.ceil(total / 1000.0 - 1e-9)


if __name__ == "__main__":
    sys.exit(main())
